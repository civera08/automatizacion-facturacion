# generar_reporte.py
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from calcular_kpis import unificar_kpis

def generar_excel_reporte(nombre_archivo=r'C:\Ruta\Al\Proyecto\seguimiento_kpis\archivo.xlsx'):
    df = unificar_kpis()

    wb = Workbook()
    ws = wb.active
    ws.title = "KPIs Empleados"

    # Encabezados con formato
    encabezados = list(df.columns)
    ws.append(encabezados)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")

    # Agregar los datos del DataFrame
    for row in dataframe_to_rows(df, index=False, header=False):
        ws.append(row)

    # Ajuste de ancho automático
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = max_length + 2

    # Guardar el archivo
    wb.save(nombre_archivo)
    print(f"✅ Reporte generado: {nombre_archivo}")

if __name__ == "__main__":
    generar_excel_reporte()